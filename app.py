from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import Body, Depends, FastAPI, Header, HTTPException, Query
from pydantic import BaseModel, Field
from sqlalchemy import (JSON, Boolean, Column, Date, DateTime, Float, ForeignKey,
                        Integer, String, Text, create_engine, func)
from sqlalchemy.orm import Session, declarative_base, relationship, sessionmaker

DATABASE_URL = "sqlite:///./ca_firm.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()

PAN_PATTERN = re.compile(r"^[A-Z]{5}[0-9]{4}[A-Z]$")
PHONE_PATTERN = re.compile(r"^[0-9]{10}$")
GSTIN_PATTERN = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$")
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class User(Base):
    __tablename__ = "users"

    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=False)
    email = Column(String, unique=True, nullable=False)
    role = Column(String, nullable=False, default="staff")  # admin, partner, manager, staff


class Client(Base):
    __tablename__ = "clients"

    id = Column(Integer, primary_key=True)
    client_code = Column(String, unique=True, nullable=False, index=True)
    client_name = Column(String, nullable=False)
    entity_type = Column(String, nullable=False)
    pan = Column(String, unique=True, nullable=False, index=True)
    gstin = Column(String, nullable=True)
    email = Column(String, nullable=False)
    phone = Column(String, nullable=False)
    assigned_partner = Column(String, nullable=True)
    assigned_manager = Column(String, nullable=True)
    status = Column(String, nullable=False, default="active")
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)

    assignments = relationship("Assignment", back_populates="client")
    billing_records = relationship("BillingRecord", back_populates="client")
    reminders = relationship("Reminder", back_populates="client")


class Assignment(Base):
    __tablename__ = "assignments"

    id = Column(Integer, primary_key=True)
    client_id = Column(Integer, ForeignKey("clients.id"), nullable=False)
    assignment_type = Column(String, nullable=False)  # GST, TDS, ITR, Audit...
    period = Column(String, nullable=False)  # e.g. FY 2025-26 / Q1-2026
    due_date = Column(Date, nullable=False)
    responsible_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    progress_percent = Column(Integer, nullable=False, default=0)
    status = Column(String, nullable=False, default="not_started")
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)

    client = relationship("Client", back_populates="assignments")


class BillingRecord(Base):
    __tablename__ = "billing_records"

    id = Column(Integer, primary_key=True)
    client_id = Column(Integer, ForeignKey("clients.id"), nullable=False)
    assignment_id = Column(Integer, ForeignKey("assignments.id"), nullable=True)
    invoice_number = Column(String, unique=True, nullable=False)
    invoice_date = Column(Date, nullable=False)
    amount = Column(Float, nullable=False)
    tax_amount = Column(Float, nullable=False, default=0)
    total_amount = Column(Float, nullable=False)
    paid_amount = Column(Float, nullable=False, default=0)
    payment_status = Column(String, nullable=False, default="unpaid")

    client = relationship("Client", back_populates="billing_records")


class Reminder(Base):
    __tablename__ = "reminders"

    id = Column(Integer, primary_key=True)
    client_id = Column(Integer, ForeignKey("clients.id"), nullable=False)
    assignment_id = Column(Integer, ForeignKey("assignments.id"), nullable=True)
    reminder_date = Column(Date, nullable=False)
    channel = Column(String, nullable=False, default="email")
    message = Column(Text, nullable=False)
    sent = Column(Boolean, nullable=False, default=False)

    client = relationship("Client", back_populates="reminders")


class UploadSession(Base):
    __tablename__ = "upload_sessions"

    id = Column(Integer, primary_key=True)
    filename = Column(String, nullable=False)
    uploaded_by = Column(Integer, ForeignKey("users.id"), nullable=True)
    uploaded_at = Column(DateTime, nullable=False, default=datetime.utcnow)
    total_rows = Column(Integer, nullable=False, default=0)
    valid_rows = Column(Integer, nullable=False, default=0)
    error_rows = Column(Integer, nullable=False, default=0)
    duplicate_rows = Column(Integer, nullable=False, default=0)
    status = Column(String, nullable=False, default="previewed")  # previewed/committed
    staged_valid_rows = Column(JSON, nullable=False, default=list)


class UploadError(Base):
    __tablename__ = "upload_errors"

    id = Column(Integer, primary_key=True)
    session_id = Column(Integer, ForeignKey("upload_sessions.id"), nullable=False)
    row_number = Column(Integer, nullable=False)
    column_name = Column(String, nullable=True)
    error_code = Column(String, nullable=False)
    message = Column(String, nullable=False)


class AuditLog(Base):
    __tablename__ = "audit_logs"

    id = Column(Integer, primary_key=True)
    entity = Column(String, nullable=False)
    entity_id = Column(String, nullable=False)
    action = Column(String, nullable=False)
    actor_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    event_ts = Column(DateTime, nullable=False, default=datetime.utcnow)
    details = Column(JSON, nullable=True)


Base.metadata.create_all(bind=engine)

app = FastAPI(title="CA Firm Management API", version="1.0.0")


class UserIn(BaseModel):
    name: str
    email: str
    role: str = Field(pattern="^(admin|partner|manager|staff)$")


class ClientIn(BaseModel):
    client_code: str
    client_name: str
    entity_type: str
    pan: str
    gstin: str | None = None
    email: str
    phone: str
    assigned_partner: str | None = None
    assigned_manager: str | None = None
    status: str = Field(default="active", pattern="^(active|inactive)$")


class AssignmentIn(BaseModel):
    client_id: int
    assignment_type: str
    period: str
    due_date: date
    responsible_user_id: int | None = None
    notes: str | None = None


class AssignmentProgressIn(BaseModel):
    progress_percent: int = Field(ge=0, le=100)
    status: str = Field(pattern="^(not_started|in_progress|completed|on_hold)$")


class BillingIn(BaseModel):
    client_id: int
    assignment_id: int | None = None
    invoice_number: str
    invoice_date: date
    amount: float = Field(gt=0)
    tax_amount: float = Field(ge=0)
    paid_amount: float = Field(default=0, ge=0)


class ReminderIn(BaseModel):
    client_id: int
    assignment_id: int | None = None
    reminder_date: date
    channel: str = Field(pattern="^(email|sms|whatsapp|call)$")
    message: str


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def ensure_default_admin() -> None:
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.role == "admin").first()
        if not admin:
            db.add(User(name="Default Admin", email="admin@example.com", role="admin"))
            db.commit()
    finally:
        db.close()


ensure_default_admin()


def get_actor(x_user_id: int | None = Header(default=None), db: Session = Depends(get_db)) -> User | None:
    if x_user_id is None:
        return None
    return db.get(User, x_user_id)


def log_audit(db: Session, entity: str, entity_id: str, action: str, actor: User | None, details: dict[str, Any]) -> None:
    db.add(
        AuditLog(
            entity=entity,
            entity_id=entity_id,
            action=action,
            actor_user_id=None if actor is None else actor.id,
            details=details,
        )
    )


def normalize_client_row(raw: dict[str, Any], row_number: int, seen_codes: set[str], seen_pan: set[str], db: Session) -> tuple[dict[str, Any] | None, list[tuple[str, str]]]:
    errors: list[tuple[str, str]] = []
    row = {k.lower().strip(): ("" if raw[k] is None else str(raw[k]).strip()) for k in raw.keys()}

    required = ["client_code", "client_name", "entity_type", "pan", "email", "phone", "status"]
    for key in required:
        if not row.get(key):
            errors.append((key, "required"))

    if row.get("pan") and not PAN_PATTERN.match(row["pan"].upper()):
        errors.append(("pan", "invalid_pan_format"))

    if row.get("phone") and not PHONE_PATTERN.match(row["phone"]):
        errors.append(("phone", "invalid_phone_format"))

    if row.get("gstin") and not GSTIN_PATTERN.match(row["gstin"].upper()):
        errors.append(("gstin", "invalid_gstin_format"))

    if row.get("status") and row["status"].lower() not in {"active", "inactive"}:
        errors.append(("status", "invalid_status"))

    code = row.get("client_code", "")
    pan = row.get("pan", "").upper()
    if code and code in seen_codes:
        errors.append(("client_code", "duplicate_in_file"))
    if pan and pan in seen_pan:
        errors.append(("pan", "duplicate_in_file"))

    if code and db.query(Client.id).filter(Client.client_code == code).first():
        errors.append(("client_code", "duplicate_in_database"))
    if pan and db.query(Client.id).filter(func.upper(Client.pan) == pan).first():
        errors.append(("pan", "duplicate_in_database"))

    seen_codes.add(code)
    seen_pan.add(pan)

    if errors:
        return None, errors

    normalized = {
        "client_code": code,
        "client_name": row["client_name"],
        "entity_type": row["entity_type"],
        "pan": pan,
        "gstin": row.get("gstin", "").upper() or None,
        "email": row["email"],
        "phone": row["phone"],
        "assigned_partner": row.get("assigned_partner") or None,
        "assigned_manager": row.get("assigned_manager") or None,
        "status": row["status"].lower(),
    }
    return normalized, []


@app.post("/users")
def create_user(payload: UserIn, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    if actor is None or actor.role != "admin":
        raise HTTPException(status_code=403, detail="only admin can create users")
    if not EMAIL_PATTERN.match(payload.email):
        raise HTTPException(status_code=400, detail="invalid email format")
    user = User(name=payload.name, email=payload.email, role=payload.role)
    db.add(user)
    db.commit()
    db.refresh(user)
    log_audit(db, "user", str(user.id), "create", actor, payload.model_dump())
    db.commit()
    return {"id": user.id, "name": user.name, "email": user.email, "role": user.role}


@app.post("/clients")
def create_client(payload: ClientIn, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    if not PAN_PATTERN.match(payload.pan.upper()):
        raise HTTPException(status_code=400, detail="invalid PAN format")
    if not PHONE_PATTERN.match(payload.phone):
        raise HTTPException(status_code=400, detail="invalid phone format")
    if not EMAIL_PATTERN.match(payload.email):
        raise HTTPException(status_code=400, detail="invalid email format")
    if payload.gstin and not GSTIN_PATTERN.match(payload.gstin.upper()):
        raise HTTPException(status_code=400, detail="invalid GSTIN format")

    exists = db.query(Client.id).filter((Client.client_code == payload.client_code) | (func.upper(Client.pan) == payload.pan.upper())).first()
    if exists:
        raise HTTPException(status_code=400, detail="client_code or PAN already exists")

    client = Client(
        client_code=payload.client_code,
        client_name=payload.client_name,
        entity_type=payload.entity_type,
        pan=payload.pan.upper(),
        gstin=None if not payload.gstin else payload.gstin.upper(),
        email=payload.email,
        phone=payload.phone,
        assigned_partner=payload.assigned_partner,
        assigned_manager=payload.assigned_manager,
        status=payload.status,
    )
    db.add(client)
    db.commit()
    db.refresh(client)
    log_audit(db, "client", str(client.id), "create", actor, payload.model_dump())
    db.commit()
    return {"id": client.id, "client_code": client.client_code, "client_name": client.client_name}


@app.get("/clients")
def list_clients(status: str | None = Query(default=None), db: Session = Depends(get_db)):
    q = db.query(Client)
    if status:
        q = q.filter(Client.status == status)
    rows = q.order_by(Client.client_name.asc()).all()
    return [
        {
            "id": c.id,
            "client_code": c.client_code,
            "client_name": c.client_name,
            "pan": c.pan,
            "status": c.status,
            "assigned_partner": c.assigned_partner,
            "assigned_manager": c.assigned_manager,
        }
        for c in rows
    ]


@app.post("/assignments")
def create_assignment(payload: AssignmentIn, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    client = db.get(Client, payload.client_id)
    if not client:
        raise HTTPException(status_code=404, detail="client not found")

    assignment = Assignment(
        client_id=payload.client_id,
        assignment_type=payload.assignment_type,
        period=payload.period,
        due_date=payload.due_date,
        responsible_user_id=payload.responsible_user_id,
        notes=payload.notes,
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    log_audit(db, "assignment", str(assignment.id), "create", actor, payload.model_dump())
    db.commit()
    return {"id": assignment.id, "status": assignment.status}


@app.patch("/assignments/{assignment_id}/progress")
def update_progress(assignment_id: int, payload: AssignmentProgressIn, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    assignment = db.get(Assignment, assignment_id)
    if not assignment:
        raise HTTPException(status_code=404, detail="assignment not found")
    assignment.progress_percent = payload.progress_percent
    assignment.status = payload.status
    db.commit()
    log_audit(db, "assignment", str(assignment.id), "update_progress", actor, payload.model_dump())
    db.commit()
    return {"id": assignment.id, "progress_percent": assignment.progress_percent, "status": assignment.status}


@app.post("/billing")
def create_billing(payload: BillingIn, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    if not db.get(Client, payload.client_id):
        raise HTTPException(status_code=404, detail="client not found")
    total_amount = round(payload.amount + payload.tax_amount, 2)
    if payload.paid_amount > total_amount:
        raise HTTPException(status_code=400, detail="paid amount cannot exceed invoice total")

    billing = BillingRecord(
        client_id=payload.client_id,
        assignment_id=payload.assignment_id,
        invoice_number=payload.invoice_number,
        invoice_date=payload.invoice_date,
        amount=payload.amount,
        tax_amount=payload.tax_amount,
        total_amount=total_amount,
        paid_amount=payload.paid_amount,
        payment_status="paid" if payload.paid_amount == total_amount else ("partial" if payload.paid_amount > 0 else "unpaid"),
    )
    db.add(billing)
    db.commit()
    db.refresh(billing)
    log_audit(db, "billing", str(billing.id), "create", actor, payload.model_dump())
    db.commit()
    return {"id": billing.id, "invoice_number": billing.invoice_number, "payment_status": billing.payment_status}


@app.post("/reminders")
def create_reminder(payload: ReminderIn, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    if not db.get(Client, payload.client_id):
        raise HTTPException(status_code=404, detail="client not found")
    reminder = Reminder(**payload.model_dump())
    db.add(reminder)
    db.commit()
    db.refresh(reminder)
    log_audit(db, "reminder", str(reminder.id), "create", actor, payload.model_dump())
    db.commit()
    return {"id": reminder.id, "reminder_date": reminder.reminder_date.isoformat(), "sent": reminder.sent}


@app.get("/dashboard/upcoming-dues")
def upcoming_dues(days: int = Query(default=15, ge=1, le=90), db: Session = Depends(get_db)):
    today = date.today()
    cutoff = today + timedelta(days=days)
    assignments = (
        db.query(Assignment, Client.client_name)
        .join(Client, Client.id == Assignment.client_id)
        .filter(Assignment.due_date >= today, Assignment.due_date <= cutoff, Assignment.status != "completed")
        .order_by(Assignment.due_date.asc())
        .all()
    )
    return [
        {
            "assignment_id": a.id,
            "client_name": client_name,
            "assignment_type": a.assignment_type,
            "due_date": a.due_date.isoformat(),
            "status": a.status,
            "progress_percent": a.progress_percent,
        }
        for a, client_name in assignments
    ]


@app.post("/bulk-upload/preview")
def preview_bulk_upload(
    filename: str = Query(..., description="Original filename ending with .xlsx"),
    content: bytes = Body(..., media_type="application/octet-stream"),
    db: Session = Depends(get_db),
    actor: User | None = Depends(get_actor),
):
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="only .xlsx files are supported")

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive parsing error
        raise HTTPException(status_code=400, detail=f"invalid excel file: {exc}") from exc

    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="excel file is empty")

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    expected = [
        "client_code",
        "client_name",
        "entity_type",
        "pan",
        "gstin",
        "email",
        "phone",
        "assigned_partner",
        "assigned_manager",
        "status",
    ]
    if headers != expected:
        raise HTTPException(status_code=400, detail=f"invalid template headers. expected {expected}")

    seen_codes: set[str] = set()
    seen_pans: set[str] = set()
    valid_rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []

    session = UploadSession(filename=filename, uploaded_by=None if actor is None else actor.id)
    db.add(session)
    db.commit()
    db.refresh(session)

    duplicate_rows = 0
    for idx, raw_values in enumerate(rows[1:], start=2):
        raw = {headers[i]: raw_values[i] if i < len(raw_values) else "" for i in range(len(headers))}
        normalized, errors = normalize_client_row(raw, idx, seen_codes, seen_pans, db)
        if normalized:
            valid_rows.append(normalized)
            preview_rows.append({"row_number": idx, "status": "valid", "data": normalized})
            continue

        is_duplicate = any(err_code.startswith("duplicate") for _, err_code in errors)
        if is_duplicate:
            duplicate_rows += 1

        row_errors = []
        for col, err_code in errors:
            row_errors.append({"column": col, "error_code": err_code})
            db.add(
                UploadError(
                    session_id=session.id,
                    row_number=idx,
                    column_name=col,
                    error_code=err_code,
                    message=f"{col}: {err_code}",
                )
            )
        preview_rows.append({"row_number": idx, "status": "error", "errors": row_errors})

    session.total_rows = max(len(rows) - 1, 0)
    session.valid_rows = len(valid_rows)
    session.error_rows = session.total_rows - session.valid_rows
    session.duplicate_rows = duplicate_rows
    session.staged_valid_rows = valid_rows
    db.commit()

    log_audit(
        db,
        "upload_session",
        str(session.id),
        "preview",
        actor,
        {
            "filename": filename,
            "total_rows": session.total_rows,
            "valid_rows": session.valid_rows,
            "error_rows": session.error_rows,
            "duplicate_rows": session.duplicate_rows,
        },
    )
    db.commit()

    return {
        "session_id": session.id,
        "filename": filename,
        "summary": {
            "total_rows": session.total_rows,
            "valid_rows": session.valid_rows,
            "error_rows": session.error_rows,
            "duplicate_rows": session.duplicate_rows,
        },
        "preview": preview_rows,
    }


@app.post("/bulk-upload/commit/{session_id}")
def commit_bulk_upload(session_id: int, db: Session = Depends(get_db), actor: User | None = Depends(get_actor)):
    session = db.get(UploadSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="upload session not found")
    if session.status == "committed":
        raise HTTPException(status_code=400, detail="session already committed")

    inserted = 0
    skipped = 0
    for row in session.staged_valid_rows or []:
        exists = db.query(Client.id).filter((Client.client_code == row["client_code"]) | (func.upper(Client.pan) == row["pan"])).first()
        if exists:
            skipped += 1
            continue
        db.add(Client(**row))
        inserted += 1

    session.status = "committed"
    db.commit()

    log_audit(
        db,
        "upload_session",
        str(session.id),
        "commit",
        actor,
        {"inserted": inserted, "skipped": skipped, "valid_rows": session.valid_rows},
    )
    db.commit()

    return {"session_id": session.id, "status": session.status, "inserted": inserted, "skipped": skipped}


@app.get("/audit-logs")
def list_audit_logs(limit: int = Query(default=100, ge=1, le=500), db: Session = Depends(get_db)):
    rows = db.query(AuditLog).order_by(AuditLog.event_ts.desc()).limit(limit).all()
    return [
        {
            "id": r.id,
            "entity": r.entity,
            "entity_id": r.entity_id,
            "action": r.action,
            "actor_user_id": r.actor_user_id,
            "event_ts": r.event_ts.isoformat(),
            "details": r.details,
        }
        for r in rows
    ]
